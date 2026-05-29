import base64
import json
import os
import uuid

import click
from collections import OrderedDict
from datetime import datetime
from functools import wraps
from hmac import compare_digest
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional

import requests
from dotenv import load_dotenv
from flask import (
    Flask,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

try:
    import boto3
except ImportError:
    boto3 = None


load_dotenv()

APP_NAME = "Preenchimento de Dados - Franqueados J&T"

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = BASE_DIR / "uploads"
SUBMISSIONS_FILE = DATA_DIR / "submissions.json"

DATA_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "webp"}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret-key-jt")


@app.cli.command("generate-admin-hash")
@click.argument("password", required=False)
def generate_admin_hash_command(password: Optional[str] = None):
    """Gera um hash seguro para usar em ADMIN_PASSWORD_HASH."""
    password = password or os.getenv("ADMIN_PASSWORD")
    if not password:
        raise click.UsageError("Informe a senha: flask --app app generate-admin-hash sua-senha")
    click.echo(generate_password_hash(password))


UF_OPTIONS = [
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA",
    "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN",
    "RS", "RO", "RR", "SC", "SP", "SE", "TO",
]

BANCO_OPTIONS = [
    "001 - Banco do Brasil",
    "033 - Santander",
    "104 - Caixa Econômica Federal",
    "237 - Bradesco",
    "341 - Itaú",
    "745 - Citibank",
    "399 - HSBC",
    "756 - Bancoob",
    "136 - Unicred",
    "290 - PagBank (PagSeguro)",
    "260 - Nubank",
    "323 - Mercado Pago",
    "380 - PicPay",
    "077 - Inter",
    "212 - Original",
    "655 - Votorantim (BV)",
    "739 - Banco Cetelem",
    "741 - Banrisul",
    "748 - Sicredi",
    "003 - Banco da Amazônia",
    "004 - Banco do Nordeste",
]

TIPO_PIX_OPTIONS = [
    "Celular",
    "Email",
    "CPF",
    "CNPJ",
    "Chave Aleatória",
    "Não possui chave pix cadastrada",
]

REGIAO_REDE_OPTIONS = [
    "MG",
    "SP",
]


FIELD_SECTIONS = [
    {
        "key": "informacoes_basicas",
        "title": "Informações básicas",
        "description": "Dados principais da empresa, localização, documentos fiscais e contato.",
        "fields": [
            {
                "name": "nome_empresa",
                "label": "Nome da empresa",
                "type": "text",
                "required": True,
                "wide": True,
            },
            {
                "name": "cnpj",
                "label": "CNPJ",
                "type": "text",
                "required": True,
                "mask": "cnpj",
            },
            {
                "name": "abreviacao_empresa",
                "label": "Nome Fantasia",
                "type": "text",
                "required": True,
            },
            {
                "name": "endereco_fornecedor",
                "label": "Endereço do franqueado",
                "type": "text",
                "required": True,
                "wide": True,
            },
            {
                "name": "uf_pertencente",
                "label": "UF pertencente",
                "type": "select",
                "required": True,
                "options": UF_OPTIONS,
            },
            {
                "name": "inscricao_estadual",
                "label": "Inscrição Estadual (I.E.)",
                "type": "text",
                "required": True,
            },
            {
                "name": "cidade_origem",
                "label": "Cidade Franquia",
                "type": "text",
                "required": True,
            },
            {
                "name": "inscricao_municipal",
                "label": "Inscrição Municipal (I.M.)",
                "type": "text",
                "required": True,
            },
            {
                "name": "codigo_postal",
                "label": "CEP",
                "type": "text",
                "required": True,
                "mask": "cep",
            },
            {
                "name": "email",
                "label": "E-mail",
                "type": "email",
                "required": True,
            },
            {
                "name": "telefone",
                "label": "DDD/Telefone",
                "type": "text",
                "required": True,
                "mask": "phone",
            },
        ],
    },
    {
        "key": "registro_cliente",
        "title": "Responsável Financeiro",
        "description": "Informações do responsável financeiro do cliente.",
        "fields": [
            {
                "name": "nome_responsavel_financeiro",
                "label": "Nome do responsável financeiro",
                "type": "text",
                "required": True,
                "wide": True,
            },
            {
                "name": "telefone_responsavel_financeiro",
                "label": "Telefone do responsável financeiro",
                "type": "text",
                "required": True,
                "mask": "phone",
            },
            {
                "name": "email_responsavel_financeiro",
                "label": "E-mail do responsável financeiro",
                "type": "email",
                "required": True,
            },
        ],
    },
    {
        "key": "testemunha_contratada",
        "title": "3 - Testemunha Contratada (Contrato)",
        "description": "Indicar uma testemunha vinculada à empresa contratada, que não seja sócia da pessoa jurídica.",
        "description_html": """
            <p>Indicar uma testemunha vinculada à empresa contratada, que não seja sócia da pessoa jurídica.</p>
            <p class=\"section-description-strong\">Pontos importantes:</p>
            <ul class=\"section-description-list section-description-strong\">
              <li>A testemunha não poderá ser nenhum dos sócios constantes no contrato social.</li>
              <li>O email da testemunha precisa conter o nome dela (ex: adrianoxxxxxx@gmail.com).</li>
            </ul>
        """,
        "fields": [
            {
                "name": "testemunha_nome_completo",
                "label": "Nome Completo",
                "type": "text",
                "required": True,
                "wide": True,
            },
            {
                "name": "testemunha_cpf",
                "label": "CPF",
                "type": "text",
                "required": True,
                "mask": "cpf",
            },
            {
                "name": "testemunha_email",
                "label": "Email",
                "type": "email",
                "required": True,
            },
            {
                "name": "testemunha_telefone_contato",
                "label": "Telefone para contato",
                "type": "text",
                "required": True,
                "mask": "phone",
            },
        ],
    },
    {
        "key": "franqueado_ativado",
        "title": "Responsável pela Franquia",
        "description": "Dados operacionais, atendimento e rota dedicada.",
        "fields": [
            {
                "name": "responsavel",
                "label": "Responsável",
                "type": "text",
                "required": True,
            },
            {
                "name": "telefone_contato",
                "label": "Telefone de contato",
                "type": "text",
                "required": True,
                "mask": "phone",
            },
            {
                "name": "tipo_abertura",
                "label": "Tipo de abertura",
                "type": "hidden",
                "value": "Transferência de Próprio para Terceiro",
                "required": False,
            },
            {
                "name": "modo_operacao",
                "label": "Modo de operação",
                "type": "hidden",
                "value": "Parceiro",
                "required": False,
            },
            {
                "name": "contato_atendimento_cliente",
                "label": "Contato do atendimento ao cliente",
                "type": "text",
                "required": True,
                "mask": "phone",
            },
            {
                "name": "atendimento_cliente",
                "label": "Atendimento ao cliente",
                "type": "text",
                "required": True,
            },
            {
                "name": "rota_dedicada",
                "label": "Rota dedicada",
                "type": "hidden",
                "value": "Não",
                "required": False,
            },
        ],
    },
    {
        "key": "criacao_feishu",
        "title": "Criação de Feishu",
        "description": "Dados pessoais necessários para criação do acesso no Feishu.",
        "fields": [
            {
                "name": "feishu_nome",
                "label": "Nome",
                "type": "text",
                "required": True,
                "wide": True,
            },
            {
                "name": "feishu_cpf",
                "label": "CPF",
                "type": "text",
                "required": True,
                "mask": "cpf",
            },
            {
                "name": "feishu_telefone_celular",
                "label": "DDD + Telefone Celular",
                "type": "text",
                "required": True,
                "mask": "phone",
            },
        ],
    },
    {
        "key": "fornecedor_registrado",
        "title": "Dados Bancários",
        "description": "Dados bancários, titular da conta e chave PIX para registro da empresa.",
        "fields": [
            {
                "name": "nome_proprietario",
                "label": "Titular da Conta",
                "type": "text",
                "required": True,
                "wide": True,
            },
            {
                "name": "numero_banco",
                "label": "Número do banco",
                "type": "select",
                "required": True,
                "options": BANCO_OPTIONS,
                "wide": True,
            },
            {
                "name": "codigo_banco_outro",
                "label": "Código do banco não listado",
                "type": "text",
                "required": False,
                "wide": True,
                "show_if_filled": True,
                "placeholder": "Preencha somente se o banco não estiver listado acima",
            },
            {
                "name": "numero_agencia",
                "label": "Número da agência",
                "type": "text",
                "required": True,
            },
            {
                "name": "conta",
                "label": "Conta",
                "type": "text",
                "required": True,
            },
            {
                "name": "tipos_contas",
                "label": "Tipo de conta",
                "type": "hidden",
                "value": "Corrente",
                "required": False,
            },
            {
                "name": "tipo_pix",
                "label": "Tipo do Pix",
                "type": "select",
                "required": True,
                "options": TIPO_PIX_OPTIONS,
            },
            {
                "name": "chave_pix",
                "label": "Chave Pix",
                "type": "text",
                "required": False,
                "wide": True,
            },
        ],
    },
]


UPLOAD_FIELDS = {
    "anexo_cnpj": "Cartão CNPJ",
    "anexo_cartao_social_ccmei": "Cartão Social / CCMEI",
    "anexo_sintegra_card": "Comprovante Sintegra",
}

UPLOAD_LINKS = {
    "anexo_cnpj": "https://solucoes.receita.fazenda.gov.br/Servicos/cnpjreva/",
    "anexo_sintegra_card": "http://www.sintegra.gov.br/",
}



def safe_date_path(dt: datetime) -> Path:
    return Path(dt.strftime("%Y")) / dt.strftime("%m") / dt.strftime("%d")


def display_date(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y")


def display_time(dt: datetime) -> str:
    return dt.strftime("%H:%M:%S")


def allowed_file(filename: str) -> bool:
    if "." not in filename:
        return False

    extension = filename.rsplit(".", 1)[1].lower()
    return extension in ALLOWED_EXTENSIONS


def load_submissions() -> List[Dict]:
    if not SUBMISSIONS_FILE.exists():
        return []

    try:
        with open(SUBMISSIONS_FILE, "r", encoding="utf-8") as file:
            return json.load(file)
    except json.JSONDecodeError:
        return []


def save_submissions(submissions: List[Dict]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    with open(SUBMISSIONS_FILE, "w", encoding="utf-8") as file:
        json.dump(submissions, file, ensure_ascii=False, indent=2)


def github_configured() -> bool:
    return bool(
        os.getenv("GITHUB_TOKEN")
        and os.getenv("GITHUB_REPO")
    )


def github_submission_path(submission: Dict) -> str:
    data_path = os.getenv("GITHUB_DATA_PATH", "submissions")
    filename = f"{submission['created_at'].replace(':', '-')}_{submission['id']}.json"
    return f"{data_path}/{submission['created_year']}/{submission['created_month']}/{submission['created_day']}/{filename}"


def r2_configured() -> bool:
    return bool(
        boto3
        and os.getenv("CLOUDFLARE_R2_ACCOUNT_ID")
        and os.getenv("CLOUDFLARE_R2_ACCESS_KEY_ID")
        and os.getenv("CLOUDFLARE_R2_SECRET_ACCESS_KEY")
        and os.getenv("CLOUDFLARE_R2_BUCKET")
    )


def push_submission_to_github(submission: Dict) -> bool:
    if not github_configured():
        return False

    token = os.getenv("GITHUB_TOKEN", "")
    repo = os.getenv("GITHUB_REPO", "")
    branch = os.getenv("GITHUB_BRANCH", "main")
    path = github_submission_path(submission)

    api_url = f"https://api.github.com/repos/{repo}/contents/{path}"

    content = json.dumps(submission, ensure_ascii=False, indent=2)
    encoded_content = base64.b64encode(content.encode("utf-8")).decode("utf-8")

    payload = {
        "message": f"Novo preenchimento de dados J&T - {submission['id']}",
        "content": encoded_content,
        "branch": branch,
    }

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
    }

    try:
        response = requests.put(api_url, headers=headers, json=payload, timeout=25)
        return response.status_code in {200, 201}
    except requests.RequestException:
        return False


def delete_submission_from_github(submission: Dict) -> bool:
    if not github_configured():
        return False

    token = os.getenv("GITHUB_TOKEN", "")
    repo = os.getenv("GITHUB_REPO", "")
    branch = os.getenv("GITHUB_BRANCH", "main")
    path = github_submission_path(submission)
    api_url = f"https://api.github.com/repos/{repo}/contents/{path}"

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
    }

    try:
        get_response = requests.get(api_url, headers=headers, params={"ref": branch}, timeout=25)

        if get_response.status_code != 200:
            return False

        sha = get_response.json().get("sha")

        if not sha:
            return False

        payload = {
            "message": f"Remove preenchimento de dados J&T - {submission['id']}",
            "sha": sha,
            "branch": branch,
        }

        delete_response = requests.delete(api_url, headers=headers, json=payload, timeout=25)
        return delete_response.status_code == 200
    except requests.RequestException:
        return False


def delete_upload_r2(key: str) -> bool:
    if not key or not r2_configured():
        return False

    boto3_module = boto3

    if boto3_module is None:
        return False

    account_id = os.getenv("CLOUDFLARE_R2_ACCOUNT_ID")
    bucket = os.getenv("CLOUDFLARE_R2_BUCKET")

    if not account_id or not bucket:
        return False

    endpoint_url = f"https://{account_id}.r2.cloudflarestorage.com"

    try:
        client = boto3_module.client(
            "s3",
            endpoint_url=endpoint_url,
            aws_access_key_id=os.getenv("CLOUDFLARE_R2_ACCESS_KEY_ID"),
            aws_secret_access_key=os.getenv("CLOUDFLARE_R2_SECRET_ACCESS_KEY"),
            region_name="auto",
        )

        client.delete_object(Bucket=bucket, Key=key)
        return True
    except Exception:
        return False


def delete_upload_local(path: str) -> bool:
    if not path:
        return False

    try:
        full_path = (UPLOAD_DIR / path).resolve()
        upload_root = UPLOAD_DIR.resolve()

        if not str(full_path).startswith(str(upload_root)):
            return False

        if full_path.exists() and full_path.is_file():
            full_path.unlink()
            return True
    except OSError:
        return False

    return False


def delete_attachment_from_storage(attachment: Dict) -> bool:
    storage = attachment.get("storage", "")
    path = attachment.get("path", "")

    if storage == "cloudflare_r2":
        return delete_upload_r2(path)

    return delete_upload_local(path)


def delete_submission_attachments(submission: Dict) -> int:
    deleted_count = 0

    for attachment in submission.get("attachments", {}).values():
        if delete_attachment_from_storage(attachment):
            deleted_count += 1

    return deleted_count


def save_upload_local(file_storage, submission_id: str, dt: datetime, upload_key: str) -> Dict:
    original_name = secure_filename(file_storage.filename or "arquivo")

    if "." not in original_name:
        extension = "bin"
    else:
        extension = original_name.rsplit(".", 1)[1].lower()

    filename = f"{submission_id}_{upload_key}.{extension}"
    folder = UPLOAD_DIR / safe_date_path(dt)
    folder.mkdir(parents=True, exist_ok=True)

    file_path = folder / filename
    file_storage.save(file_path)

    relative = f"{safe_date_path(dt).as_posix()}/{filename}"

    return {
        "storage": "local",
        "label": UPLOAD_FIELDS[upload_key],
        "original_name": original_name,
        "filename": filename,
        "path": relative,
        "url": url_for("uploaded_file", filepath=relative, _external=False),
        "download_url": url_for("download_uploaded_file", filepath=relative, _external=False),
    }


def save_upload_r2(file_storage, submission_id: str, dt: datetime, upload_key: str) -> Optional[Dict]:
    if not r2_configured():
        return None

    boto3_module = boto3

    if boto3_module is None:
        return None

    original_name = secure_filename(file_storage.filename or "arquivo")

    if "." not in original_name:
        return None

    extension = original_name.rsplit(".", 1)[1].lower()
    key = f"cadastros/{safe_date_path(dt).as_posix()}/{submission_id}/{upload_key}.{extension}"

    account_id = os.getenv("CLOUDFLARE_R2_ACCOUNT_ID")
    bucket = os.getenv("CLOUDFLARE_R2_BUCKET")

    if not account_id or not bucket:
        return None

    endpoint_url = f"https://{account_id}.r2.cloudflarestorage.com"

    client = boto3_module.client(
        "s3",
        endpoint_url=endpoint_url,
        aws_access_key_id=os.getenv("CLOUDFLARE_R2_ACCESS_KEY_ID"),
        aws_secret_access_key=os.getenv("CLOUDFLARE_R2_SECRET_ACCESS_KEY"),
        region_name="auto",
    )

    file_storage.stream.seek(0)

    client.upload_fileobj(
        file_storage.stream,
        bucket,
        key,
        ExtraArgs={"ContentType": file_storage.mimetype or "application/octet-stream"},
    )

    public_base = os.getenv("CLOUDFLARE_R2_PUBLIC_URL", "").rstrip("/")

    public_url = f"{public_base}/{key}" if public_base else ""

    return {
        "storage": "cloudflare_r2",
        "label": UPLOAD_FIELDS[upload_key],
        "original_name": original_name,
        "filename": key.split("/")[-1],
        "path": key,
        "url": public_url,
        "download_url": public_url,
    }


def save_upload(file_storage, submission_id: str, dt: datetime, upload_key: str) -> Dict:
    try:
        r2_result = save_upload_r2(file_storage, submission_id, dt, upload_key)

        if r2_result:
            return r2_result
    except Exception:
        try:
            file_storage.stream.seek(0)
        except Exception:
            pass

    return save_upload_local(file_storage, submission_id, dt, upload_key)


def collect_form_data() -> Dict:
    data = {}

    for section in FIELD_SECTIONS:
        for field in section["fields"]:
            field_name = field["name"]

            if field.get("type") == "hidden":
                data[field_name] = field.get("value", "").strip()
            else:
                data[field_name] = request.form.get(field_name, "").strip()

    data["tipo_abertura"] = "Transferência de Próprio para Terceiro"
    data["modo_operacao"] = "Parceiro"
    data["rota_dedicada"] = "Não"
    data["tipos_contas"] = "Corrente"

    return data


def validate_required_fields(data: Dict) -> List[str]:
    missing = []

    for section in FIELD_SECTIONS:
        for field in section["fields"]:
            if field.get("type") == "hidden":
                continue

            if field.get("required") and not data.get(field["name"]):
                missing.append(field["label"])

    return missing


def validate_uploads() -> List[str]:
    errors = []

    for upload_key, upload_label in UPLOAD_FIELDS.items():
        file_storage = request.files.get(upload_key)

        if not file_storage or not file_storage.filename:
            errors.append(f"Anexo obrigatório: {upload_label}")
            continue

        if not allowed_file(file_storage.filename):
            errors.append(f"Formato inválido em {upload_label}. Use PDF, PNG, JPG, JPEG ou WEBP.")

    return errors


def find_submission(submission_id: str) -> Optional[Dict]:
    submissions = load_submissions()

    for submission in submissions:
        if submission.get("id") == submission_id:
            return submission

    return None


def group_submissions_by_date(submissions: List[Dict]) -> OrderedDict:
    ordered = sorted(
        submissions,
        key=lambda item: item.get("created_at", ""),
        reverse=True,
    )

    grouped = OrderedDict()

    for item in ordered:
        date_key = item.get("created_date", "Sem data")

        if date_key not in grouped:
            grouped[date_key] = []

        grouped[date_key].append(item)

    for date_key in grouped:
        grouped[date_key] = sorted(
            grouped[date_key],
            key=lambda item: item.get("created_at", ""),
        )

    return grouped


def parse_filter_date(value: str, end_of_day: bool = False) -> Optional[datetime]:
    if not value:
        return None

    try:
        parsed = datetime.strptime(value, "%Y-%m-%d")

        if end_of_day:
            return parsed.replace(hour=23, minute=59, second=59)

        return parsed.replace(hour=0, minute=0, second=0)
    except ValueError:
        return None


def filter_submissions_by_period(submissions: List[Dict], start_date: str, end_date: str) -> List[Dict]:
    start_dt = parse_filter_date(start_date, end_of_day=False)
    end_dt = parse_filter_date(end_date, end_of_day=True)

    filtered = []

    for submission in submissions:
        created_at = submission.get("created_at", "")

        try:
            created_dt = datetime.fromisoformat(created_at)
        except ValueError:
            continue

        if start_dt and created_dt < start_dt:
            continue

        if end_dt and created_dt > end_dt:
            continue

        filtered.append(submission)

    return sorted(filtered, key=lambda item: item.get("created_at", ""))


def should_include_field_for_export(field: Dict, submissions: Optional[List[Dict]] = None) -> bool:
    if not field.get("show_if_filled"):
        return True

    if submissions is None:
        return True

    field_name = field.get("name", "")

    return any(
        str(submission.get("data", {}).get(field_name, "")).strip()
        for submission in submissions
    )


def get_export_headers(submissions: Optional[List[Dict]] = None) -> List[Dict]:
    headers = [
        {"section": "Controle", "label": "ID", "name": "__id"},
        {"section": "Controle", "label": "Data de preenchimento", "name": "__created_date"},
        {"section": "Controle", "label": "Horário de preenchimento", "name": "__created_time"},
    ]

    for section in FIELD_SECTIONS:
        for field in section["fields"]:
            if not should_include_field_for_export(field, submissions):
                continue

            headers.append({
                "section": section["title"],
                "label": field["label"],
                "name": field["name"],
            })

    for upload_key, upload_label in UPLOAD_FIELDS.items():
        headers.append({
            "section": "Anexos",
            "label": f"{upload_label} - Nome do arquivo",
            "name": f"__attachment_name_{upload_key}",
        })
        headers.append({
            "section": "Anexos",
            "label": f"{upload_label} - Link",
            "name": f"__attachment_url_{upload_key}",
        })

    return headers


def value_for_export(submission: Dict, header: Dict) -> str:
    name = header["name"]

    if name == "__id":
        return submission.get("id", "")

    if name == "__created_date":
        return submission.get("created_date", "")

    if name == "__created_time":
        return submission.get("created_time", "")

    if name.startswith("__attachment_name_"):
        upload_key = name.replace("__attachment_name_", "")
        attachment = submission.get("attachments", {}).get(upload_key, {})
        return attachment.get("original_name", "")

    if name.startswith("__attachment_url_"):
        upload_key = name.replace("__attachment_url_", "")
        attachment = submission.get("attachments", {}).get(upload_key, {})
        return attachment.get("download_url") or attachment.get("url", "")

    return submission.get("data", {}).get(name, "")


def build_xlsx(submissions: List[Dict], start_date: str = "", end_date: str = "") -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active

    # Corrige o alerta do Pylance: Workbook.active pode ser tipado como None.
    if sheet is None:
        sheet = workbook.create_sheet("Cadastros", 0)

    sheet.title = "Cadastros"

    headers = get_export_headers(submissions)

    title = "Relatório - Preenchimento de Dados - Franqueados J&T"
    if start_date or end_date:
        title += f" | Período: {start_date or 'início'} até {end_date or 'hoje'}"

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=15, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="E60012")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_index, header in enumerate(headers, start=1):
        section_cell = sheet.cell(row=2, column=col_index, value=header["section"])
        section_cell.font = Font(bold=True, color="FFFFFF")
        section_cell.fill = PatternFill("solid", fgColor="222222")
        section_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        label_cell = sheet.cell(row=3, column=col_index, value=header["label"])
        label_cell.font = Font(bold=True, color="FFFFFF")
        label_cell.fill = PatternFill("solid", fgColor="E60012")
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, submission in enumerate(submissions, start=4):
        for col_index, header in enumerate(headers, start=1):
            value = value_for_export(submission, header)
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{max(4, len(submissions) + 3)}"

    for column_index in range(1, len(headers) + 1):
        column_letter = get_column_letter(column_index)
        max_length = 12

        for row in range(1, min(sheet.max_row, 40) + 1):
            value = sheet.cell(row=row, column=column_index).value
            if value:
                max_length = max(max_length, len(str(value)))

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 42)

    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 34

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def admin_required(route_function):
    @wraps(route_function)
    def wrapper(*args, **kwargs):
        if not session.get("admin_logged"):
            return redirect(url_for("admin_login"))

        return route_function(*args, **kwargs)

    return wrapper


@app.context_processor
def inject_globals():
    return {
        "APP_NAME": APP_NAME,
        "FIELD_SECTIONS": FIELD_SECTIONS,
        "UPLOAD_FIELDS": UPLOAD_FIELDS,
        "UPLOAD_LINKS": UPLOAD_LINKS,
    }


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/submit", methods=["POST"])
def submit():
    data = collect_form_data()

    missing_fields = validate_required_fields(data)
    upload_errors = validate_uploads()

    if missing_fields or upload_errors:
        for item in missing_fields:
            flash(f"Campo obrigatório não preenchido: {item}", "error")

        for item in upload_errors:
            flash(item, "error")

        return redirect(url_for("index"))

    now = datetime.now()
    submission_id = uuid.uuid4().hex[:12].upper()

    attachments = {}

    for upload_key in UPLOAD_FIELDS:
        file_storage = request.files.get(upload_key)

        if file_storage and file_storage.filename:
            attachments[upload_key] = save_upload(file_storage, submission_id, now, upload_key)

    submission = {
        "id": submission_id,
        "created_at": now.isoformat(timespec="seconds"),
        "created_date": display_date(now),
        "created_time": display_time(now),
        "created_year": now.strftime("%Y"),
        "created_month": now.strftime("%m"),
        "created_day": now.strftime("%d"),
        "data": data,
        "attachments": attachments,
        "storage_status": {
            "github": False,
            "cloudflare_r2": any(
                attachment.get("storage") == "cloudflare_r2"
                for attachment in attachments.values()
            ),
        },
    }

    submission["storage_status"]["github"] = push_submission_to_github(submission)

    submissions = load_submissions()
    submissions.append(submission)
    save_submissions(submissions)

    return redirect(url_for("success", submission_id=submission_id))


@app.route("/success/<submission_id>")
def success(submission_id: str):
    submission = find_submission(submission_id)

    if not submission:
        abort(404)

    return render_template("success.html", submission=submission)


@app.route("/uploads/<path:filepath>")
def uploaded_file(filepath: str):
    full_path = (UPLOAD_DIR / filepath).resolve()
    upload_root = UPLOAD_DIR.resolve()

    if not str(full_path).startswith(str(upload_root)):
        abort(404)

    if not full_path.exists():
        abort(404)

    return send_from_directory(full_path.parent, full_path.name, as_attachment=False)


@app.route("/uploads/download/<path:filepath>")
def download_uploaded_file(filepath: str):
    full_path = (UPLOAD_DIR / filepath).resolve()
    upload_root = UPLOAD_DIR.resolve()

    if not str(full_path).startswith(str(upload_root)):
        abort(404)

    if not full_path.exists():
        abort(404)

    return send_from_directory(full_path.parent, full_path.name, as_attachment=True)


@app.route("/admin", methods=["GET"])
@admin_required
def admin_panel():
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    submissions = load_submissions()
    filtered = filter_submissions_by_period(submissions, start_date, end_date) if (start_date or end_date) else submissions
    grouped = group_submissions_by_date(filtered)

    return render_template(
        "admin.html",
        grouped=grouped,
        total=len(submissions),
        preview_total=len(filtered),
        start_date=start_date,
        end_date=end_date,
        github_enabled=github_configured(),
        r2_enabled=r2_configured(),
    )


@app.route("/admin/report", methods=["GET"])
@admin_required
def admin_report():
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    submissions = load_submissions()
    filtered = filter_submissions_by_period(submissions, start_date, end_date) if (start_date or end_date) else submissions
    grouped = group_submissions_by_date(filtered)

    return render_template(
        "admin_report.html",
        grouped=grouped,
        total=len(filtered),
        start_date=start_date,
        end_date=end_date,
    )


@app.route("/admin/export/xlsx", methods=["GET"])
@admin_required
def admin_export_xlsx():
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    submissions = load_submissions()
    filtered = filter_submissions_by_period(submissions, start_date, end_date)

    output = build_xlsx(filtered, start_date, end_date)

    filename_start = start_date or "inicio"
    filename_end = end_date or "hoje"
    filename = f"preenchimento-franqueados-jt-{filename_start}-a-{filename_end}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/submission/<submission_id>/delete", methods=["POST"])
@admin_required
def admin_delete_submission(submission_id: str):
    submissions = load_submissions()
    submission_to_delete = None
    remaining_submissions = []

    for submission in submissions:
        if submission.get("id") == submission_id:
            submission_to_delete = submission
        else:
            remaining_submissions.append(submission)

    if not submission_to_delete:
        flash("Cadastro não encontrado ou já deletado.", "error")
        return redirect(url_for("admin_panel"))

    deleted_attachments = delete_submission_attachments(submission_to_delete)
    github_deleted = delete_submission_from_github(submission_to_delete)

    save_submissions(remaining_submissions)

    github_message = " Backup do GitHub removido." if github_deleted else ""
    flash(
        f"Cadastro {submission_id} deletado. {deleted_attachments} anexo(s) removido(s) do armazenamento disponível." + github_message,
        "success",
    )

    return redirect(url_for("admin_panel"))


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")

        expected_username = os.getenv("ADMIN_USERNAME") or os.getenv("ADMIN_USER") or "luiz_miranda"
        expected_password = os.getenv("ADMIN_PASSWORD", "admin123")
        expected_password_hash = os.getenv("ADMIN_PASSWORD_HASH", "").strip()

        valid_username = compare_digest(username, expected_username)
        if expected_password_hash:
            valid_password = check_password_hash(expected_password_hash, password)
        else:
            valid_password = compare_digest(password, expected_password)

        if valid_username and valid_password:
            session["admin_logged"] = True
            return redirect(url_for("admin_panel"))

        flash("Usuário ou senha inválidos.", "error")

    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged", None)
    return redirect(url_for("admin_login"))


@app.route("/admin/submission/<submission_id>")
@admin_required
def admin_submission_detail(submission_id: str):
    submission = find_submission(submission_id)

    if not submission:
        abort(404)

    return render_template("admin_detail.html", submission=submission)


@app.errorhandler(404)
def not_found(_error):
    return render_template("404.html"), 404


if __name__ == "__main__":
    app.run(debug=True)
